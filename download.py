import requests
from bs4 import BeautifulSoup as BS
import os.path
import time
import shutil
import sqlite3 as sq
import pdfplumber
import fitz
from os import listdir, mkdir
import re
import openpyxl

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def download_replacements(rep):
    try:
        req = requests.get("https://lmk-lipetsk.ru/main_razdel/shedule/index.php", timeout=10, verify=False)
        soup = BS(req.text, 'html.parser')
        zam = soup.find('div', class_='page-tmpl-content')
        urls = (zam.find_all('a'))
        for url_1 in urls:
            if 'занятий' not in url_1.text:
                pass
            else:
                url = 'https://www.lmk-lipetsk.ru' + url_1['href']
                r = requests.get(url, timeout=10, verify=False)
                name = 'zam/' + url.split('/')[-1].replace('%20', ' ')

                # Проверка на наличие файла
                if os.path.isfile(name):
                    break
                else:
                    rep.value = False
                    time.sleep(1)
                    print('\nЗАМЕНЫ')
                    if urls.index(url_1) == 0:  # Перенос файлов в архив
                        tree = os.path.dirname(os.path.abspath(__file__))
                        source = tree + '/zam'
                        destination = tree + '/arhiv'
                        allfiles = os.listdir(source)
                        for f in allfiles:
                            src_path = os.path.join(source, f)
                            dst_path = os.path.join(destination, f)
                            shutil.move(src_path, dst_path)

                    open(name, 'wb').write(r.content)

                    ras = []
                    with pdfplumber.open(name) as pdf:
                        page = pdf.pages[0]
                        day = page.extract_text()
                        for number_page in range(len(pdf.pages)):
                            page = pdf.pages[number_page]
                            try:
                                ras += page.extract_table()
                            except:
                                pass
                    day = day.split('\n')
                    if day[0] == '.':
                        day.pop(0)
                    day = f'{day[0]} {day[1]} {day[2]}'

                    shutil.rmtree('png_zam')
                    mkdir('png_zam')
                    doc = fitz.open(name)
                    for i in range(len(doc)):
                        page = doc.load_page(i)
                        pix = page.get_pixmap()
                        output = f"png_zam/zam{i}.png"
                        pix.save(output)
                    doc.close()

                    del ras[0:2]
                    zam = [["", "", ""]]
                    kostil = False

                    for i in range(len(ras)):
                        if kostil:
                            if 'Выходят' in ras[i][0] or 'Дистанционное' in ras[i][0]:
                                break
                            for j in ras[i]:
                                if j:
                                    zam.append([j.replace('ИСИП', 'ИСиП'), 'практика', '', '', day])
                            continue

                        if len(ras[i]) > 4:
                            if ras[i][0] is None:
                                edit = 1
                                ras[i][0] = 123
                            else:
                                edit = 0
                            while len(ras[i]) > 4:
                                ras[i].remove(None)
                            if edit:
                                ras[i][0] = None

                        bufer = []
                        if ras[i][0] == 'Практика':
                            kostil = True
                            continue
                        elif ras[i][0] is None:
                            bufer.append(zam[-1][0])
                        else:
                            bufer.append(ras[i][0])

                        if not ras[i][1]:
                            if ras[i][0]:
                                if 'курс' not in ras[i][0]:
                                    ras[i][1] = 0

                        if ras[i][2] is None:
                            ras[i][2] = ras[i - 1][2]

                        if ras[i][3] is None:
                            ras[i][3] = ras[i - 1][3]
                        if ras[i][2] == '':
                            bufer.append(ras[i][1])
                            bufer.append('')
                            bufer.append(ras[i][3])
                        else:
                            bufer.append(ras[i][1])
                            bufer.append(ras[i][2])
                            bufer.append(ras[i][3])

                        if bufer:
                            if bufer[1] == '' and bufer[3] == '':
                                continue

                        bufer.append(day)

                        if len(bufer) > 2:
                            if bufer[1] is not None:

                                try:
                                    if bufer[2] == 'ин. яз.' or bufer[2] == 'ин.яз.':
                                        with sq.connect('bd/LMK.db') as con:
                                            cur = con.cursor()
                                            cur.execute(
                                                """SELECT para FROM schedule WHERE (groupa = ?) AND (para LIKE '%Ино%') LIMIT 1""",
                                                [bufer[0], ])
                                            bufer[2] = cur.fetchall()[0][0]
                                    if bufer[3] == 'ин. яз.':
                                        with sq.connect('bd/LMK.db') as con:
                                            cur = con.cursor()
                                            cur.execute(
                                                """SELECT para FROM schedule WHERE (groupa = ?) AND (para LIKE '%Ино%') LIMIT 1""",
                                                [bufer[0], ])
                                            bufer[3] = cur.fetchall()[0][0]
                                except:
                                    pass

                                if bufer[0] != '':
                                    zam.append(bufer)
                                else:
                                    bufer[0] = zam[-1][0]
                                    zam.append(bufer)

                    while True:
                        if zam[0][0] == '':
                            zam.pop(0)
                        else:
                            break

                    for i in range(len(zam)-1):
                        zam[i][0] = zam[i][0].replace('ИСИП', 'ИСиП')

                    with sq.connect('bd/LMK.db') as con:
                        cur = con.cursor()

                        cur.execute('''SELECT text FROM days where type = "replacements_day"''')
                        last_day = cur.fetchall()
                        if last_day:
                            if last_day[0][0] != day:
                                cur.execute("""DROP TABLE IF EXISTS last_replacements""")
                                cur.execute(
                                    """CREATE TABLE last_replacements AS SELECT * FROM replacements""")
                                cur.execute('''DELETE FROM days where type = "last_replacements_day"''')
                                cur.execute('''REPLACE INTO days VALUES ("last_replacements_day", ?)''', (last_day[0][0],))

                        cur.execute("""DROP TABLE IF EXISTS replacements""")
                        cur.execute(
                            """CREATE TABLE replacements ("groupa" TEXT, "number_para" TEXT, "replacements1" TEXT, "replacements2" TEXT, "day" TEXT)""")

                        cur.executemany("""INSERT INTO replacements VALUES (?, ?, ?, ?, ?)""", zam)

                        cur.execute('''CREATE TABLE IF NOT EXISTS days ("type"	TEXT, "text"	TEXT)''')
                        cur.execute('''DELETE FROM days where type = "replacements_day"''')
                        cur.execute('''INSERT INTO days VALUES ("replacements_day", ?)''', (day, ))

                    print('Замены обновлены\n')
                    rep.value = True
                break

    except Exception as e:
        rep.value = True
        print('Ошибка в загрузке:', e)
        with sq.connect('bd/LMK.db') as con:
            cur = con.cursor()
            cur.execute(
                '''CREATE TABLE IF NOT EXISTS errors ("number" INTEGER, "error" TEXT, PRIMARY KEY("number" AUTOINCREMENT))''')
            cur.execute('''INSERT INTO errors (error) VALUES (?)''', (str(e),))
    try:
        del ras, zam
    except:
        pass


def download_schedule():
    while True:
        try:
            req = requests.get("https://lmk-lipetsk.ru/main_razdel/shedule/index.php", timeout=10, verify=False)
            soup = BS(req.text, 'html.parser')
            zam = soup.find('div', class_='page-tmpl-content')
            urls = (zam.find_all('a', target='_blank', class_="list-group-item"))

            shutil.rmtree('schedule')
            mkdir('schedule')

            for url_1 in urls:
                url = url_1['href']
                url = 'https://www.lmk-lipetsk.ru' + url
                r = requests.get(url, timeout=10, verify=False)
                name = 'schedule/' + url.split('/')[-1].replace('%20', ' ')
                open(name, 'wb').write(r.content)

            day_of_wek = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']
            result = []

            files = listdir('schedule')
            for file in files:
                if file[-4:] == '.pdf':
                    with pdfplumber.open("schedule/" + file) as pdf:
                        first_page = pdf.pages[0]
                        ras = first_page.extract_table()
                        n_group = 2

                        if ras[0][0].lower() != 'ь\nн\nе\nд' and ras[0][0].lower() != 'ьнед':
                            ras.pop(0)

                        while n_group < len(ras[0]) - 1:
                            if ras[0][n_group + 1] is None:
                                bn = ''
                                zn = ''
                                for i in range(len(ras)):
                                    if ras[i][1] == '1':
                                        bn += '\n'
                                        zn += '\n'
                                    if ras[i][n_group] != '':
                                        bn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'

                                    if ras[i][n_group + 1] is None:
                                        if ras[i][n_group] != '':
                                            zn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                    else:
                                        if ras[i][n_group + 1] != '':
                                            zn += f'{ras[i][1]} {ras[i][n_group + 1]}'.replace('\n', ' ') + '\n'
                                bn = bn.replace('  ', ' ')
                                bn = bn.split('\n\n')
                                zn = zn.split('\n\n')

                                name_group = f'{bn[0].split(" ")[-2]} {bn[0].split(" ")[-1]}'

                                bn.pop(0)
                                zn.pop(0)

                                for i in range(len(bn)):
                                    bn[i] = bn[i].split('\n')
                                    if bn[i] != ['']:
                                        for n in bn[i]:
                                            if n != '':
                                                number = n[0]
                                                n = n[1:]
                                                result.append([name_group, 'белая неделя', day_of_wek[i], number, n])
                                for i in range(len(zn)):
                                    zn[i] = zn[i].split('\n')
                                    if bn[i] != ['']:
                                        for n in zn[i]:
                                            if n != '':
                                                number = n[0]
                                                n = n[1:]
                                                result.append([name_group, 'зеленая неделя', day_of_wek[i], number, n])
                                n_group += 2

                            else:
                                bn = ''
                                for i in range(0, len(ras)):
                                    if ras[i][1] == '1':
                                        bn += '\n'
                                    if ras[i][n_group] != '':
                                        bn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                bn = bn.replace('  ', ' ')
                                bn = bn.split('\n\n')
                                name_group = f'{bn[0].split(" ")[-2]} {bn[0].split(" ")[-1]}'
                                bn.pop(0)
                                for i in range(len(bn)):
                                    bn[i] = bn[i].split('\n')
                                    if bn[i] != ['']:
                                        for n in bn[i]:
                                            if n != '':
                                                number = n[0]
                                                n = n[1:]
                                                result.append([name_group, 'белая неделя', day_of_wek[i], number, n])
                                                result.append([name_group, 'зеленая неделя', day_of_wek[i], number, n])
                                n_group += 1

                elif file[-5:] == '.xlsx':
                    ras = []
                    xlfile = openpyxl.load_workbook("schedule/" + file)
                    xlfile = xlfile.active
                    for i in range(0, xlfile.max_row):
                        bufer = []
                        for col in xlfile.iter_cols(1, xlfile.max_column):
                            bufer.append(col[i].value)
                        ras.append(bufer)

                    n_group = 2
                    while n_group < len(ras[0]) - 1:
                        if ras[0][n_group] is None:
                            break
                        if ras[0][n_group + 1] is None:
                            bn = ''
                            zn = ''
                            for i in range(len(ras)):
                                if ras[i][1] == 1:
                                    bn += '\n'
                                    zn += '\n'
                                if ras[i][n_group]:
                                    if ras[i][1]:
                                        bn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                    else:
                                        bn += f'  {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                if ras[i][n_group + 1] is None:
                                    if ras[i][n_group]:
                                        if ras[i][1]:
                                            zn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                        else:
                                            zn += f'  {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                else:
                                    if ras[i][n_group + 1]:
                                        if ras[i][1]:
                                            zn += f'{ras[i][1]} {ras[i][n_group + 1]}'.replace('\n', ' ') + '\n'
                                        else:
                                            zn += f'  {ras[i][n_group + 1]}'.replace('\n', ' ') + '\n'

                            bn = bn.replace('  ', ' ')
                            bn = bn.split('\n\n')
                            zn = zn.split('\n\n')

                            name_group = f'{bn[0].split(" ")[-2]} {bn[0].split(" ")[-1]}'

                            bn.pop(0)
                            zn.pop(0)

                            for i in range(len(bn)):
                                bn[i] = bn[i].split('\n')
                                if bn[i] != ['']:
                                    for n in bn[i]:
                                        if n != '':
                                            number = n[0]
                                            n = n[1:]
                                            result.append([name_group, 'белая неделя', day_of_wek[i], number, n])
                            for i in range(len(zn)):
                                zn[i] = zn[i].split('\n')
                                if bn[i] != ['']:
                                    for n in zn[i]:
                                        if n != '':
                                            number = n[0]
                                            n = n[1:]
                                            result.append([name_group, 'зеленая неделя', day_of_wek[i], number, n])
                            n_group += 2

                        else:
                            bn = ''
                            for i in range(len(ras)):
                                if ras[i][1] == 1:
                                    bn += '\n'
                                if ras[i][n_group]:
                                    if ras[i][1]:
                                        bn += f'{ras[i][1]} {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                                    else:
                                        bn += f'  {ras[i][n_group]}'.replace('\n', ' ') + '\n'
                            bn = bn.replace('  ', ' ')
                            bn = bn.split('\n\n')
                            name_group = f'{bn[0].split(" ")[-2]} {bn[0].split(" ")[-1]}'
                            bn.pop(0)
                            for i in range(len(bn)):
                                bn[i] = bn[i].split('\n')
                                if bn[i] != ['']:
                                    for n in bn[i]:
                                        if n != '':
                                            number = n[0]
                                            n = n[1:]
                                            result.append([name_group, 'белая неделя', day_of_wek[i], number, n])
                                            result.append([name_group, 'зеленая неделя', day_of_wek[i], number, n])
                            n_group += 1

            with sq.connect('bd/LMK.db') as con:
                cur = con.cursor()
                cur.execute("""DROP TABLE IF EXISTS schedule""")
                cur.execute("""CREATE TABLE IF NOT EXISTS schedule (
                            "groupa"	TEXT,
                            "week_tipe"	TEXT,
                            "day"	TEXT,
                            "number" TEXT,
                            "para" TEXT);""")
                cur.executemany("""INSERT INTO schedule VALUES (?, ?, ?, ?, ?)""", result)

            schedule = []
            prepods = []
            for i in result:
                i.pop(0)
                p = re.findall(r"[А-Я][а-я]{1,20} [А-Я]\. {0,2}[А-Я]\.", i[3])
                n = re.findall(r"[А-Я][а-я]{1,20} [А-Я]\. {0,2}[А-Я]\. {0,2}\([0-9]{3}\)", i[3])
                for q in n:
                    i = list(i)
                    prepod, i[3] = q.replace('  ', ' ').replace('. ', '.').replace('.(', '. (').split(' (')
                    i[3] = i[3][:-1]
                    if len(i) == 5:
                        i.pop(-1)
                    i.append(prepod)
                    if i not in schedule:
                        schedule.append(i.copy())
                        i.pop(-1)
                for prepod in p:
                    prepod = prepod.replace('  ', ' ').replace('. ', '.')
                    if (prepod,) not in prepods:
                        prepods.append((prepod,))

            with sq.connect('bd/LMK.db') as con:
                cur = con.cursor()
                cur.execute("""DROP TABLE IF EXISTS prepods""")
                cur.execute("""CREATE TABLE "prepods" (
                "id"	INTEGER UNIQUE,
                "prepod"	TEXT UNIQUE,
                PRIMARY KEY("id" AUTOINCREMENT))""")
                cur.executemany("""INSERT INTO prepods (prepod) VALUES (?)""", prepods)

                cur.execute("""DROP TABLE IF EXISTS prepods_cabinet""")
                cur.execute("""CREATE TABLE "prepods_cabinet" (
                "prepod"	INTEGER,
                "cabinet"	INTEGER,
                "para"	INTEGER,
                "day"	TEXT,
                "week_tipe"	TEXT,
                FOREIGN KEY("prepod") REFERENCES "prepods"("id"))""")
                cur.executemany("""INSERT INTO prepods_cabinet (week_tipe, day, para, cabinet, prepod) 
                VALUES (?,?,?,?,(SELECT id FROM prepods WHERE prepod = ?))""", schedule)

        except Exception as e:
            print('Ошибка в загрузке:', e)
        try:
            del ras, result, req, schedule, prepods
        except:
            pass

        time.sleep(3600)

